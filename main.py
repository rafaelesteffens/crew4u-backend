from fastapi import FastAPI, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from datetime import datetime, date, time, timedelta
import tempfile
import os
import re
import csv
import math
import traceback

app = FastAPI()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
AEROPORTOS_CSV = os.path.join(BASE_DIR, "aeroportos.csv")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
def health_check():
    return {"status": "ok", "message": "Crew 4U backend funcionando"}

@app.post("/upload-escala")
async def upload_escala(file: UploadFile = File(...), cargo: str = Form("COPILOTO")):
    suffix = os.path.splitext(file.filename or "")[1] or ".xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp:
        content = await file.read()
        temp.write(content)
        temp_path = temp.name

    try:
        aeroportos = carregar_aeroportos()
        workbook = load_workbook(temp_path, data_only=True)
        sheet = escolher_aba_escala(workbook)
        periodo = detectar_periodo_vigencia(sheet)
        raw_rows = extrair_linhas_brutas(sheet)
        events = extrair_eventos_escala(sheet, periodo, aeroportos)
        summary = calcular_resumo_inicial(events, aeroportos, cargo)

        return {
            "filename": file.filename,
            "sheet": sheet.title,
            "periodo": {
                "inicio": periodo["inicio"].strftime("%d/%m/%Y") if periodo else "",
                "fim": periodo["fim"].strftime("%d/%m/%Y") if periodo else "",
            },
            "rows": raw_rows,
            "events": events,
            "events_count": len(events),
            "summary": summary,
        }
    except Exception as exc:
        print("ERRO AO PROCESSAR ESCALA:")
        traceback.print_exc()
        return {
            "filename": file.filename,
            "sheet": "",
            "periodo": {"inicio": "", "fim": ""},
            "rows": [],
            "events": [],
            "events_count": 0,
            "summary": calcular_resumo_inicial([], {}, cargo),
            "error": str(exc),
            "detail": "Não foi possível interpretar esta escala. Veja o log do backend para detalhes.",
        }
    finally:
        try:
            os.remove(temp_path)
        except OSError:
            pass

# ============================================================
# WORKBOOK / ABA
# ============================================================

def escolher_aba_escala(workbook):
    # Preferir a primeira aba que pareça conter a escala.
    palavras = ["DUTY", "REPORT", "DEP", "ARR", "PAIRING", "ITEM", "ROSTER", "ESCALA"]
    melhor_sheet = workbook[workbook.sheetnames[0]]
    melhor_score = -1

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        score = 0
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 25), values_only=True):
            texto = " ".join(converter_valor_para_texto(c).upper() for c in row if c is not None)
            score += sum(1 for p in palavras if p in texto)
        if score > melhor_score:
            melhor_score = score
            melhor_sheet = sheet
    return melhor_sheet

# ============================================================
# AEROPORTOS
# ============================================================

def carregar_aeroportos():
    aeroportos = {}
    if not os.path.exists(AEROPORTOS_CSV):
        print("Arquivo aeroportos.csv não encontrado.")
        return aeroportos

    with open(AEROPORTOS_CSV, newline="", encoding="utf-8-sig") as arquivo:
        primeira_linha = arquivo.readline()
        arquivo.seek(0)
        tem_cabecalho = any(x in primeira_linha.upper() for x in ["IATA", "ICAO", "LATITUDE"])

        if tem_cabecalho:
            leitor = csv.DictReader(arquivo)
            for linha in leitor:
                iata = limpar_texto(linha.get("IATA", ""))
                icao = limpar_texto(linha.get("ICAO", ""))
                nome = limpar_texto(linha.get("NOME", ""))
                latitude = converter_numero_decimal(linha.get("LATITUDE", ""))
                longitude = converter_numero_decimal(linha.get("LONGITUDE", ""))
                pais = limpar_texto(linha.get("PAIS", "")) or detectar_pais_por_icao(icao)
                grupo_diaria = limpar_texto(linha.get("GRUPO_DIARIA", "")) or detectar_grupo_diaria_por_icao(icao)
                adicionar_aeroporto(aeroportos, iata, icao, nome, latitude, longitude, pais, grupo_diaria)
        else:
            leitor = csv.reader(arquivo)
            for linha in leitor:
                if len(linha) < 5:
                    continue
                iata = limpar_texto(linha[0])
                icao = limpar_texto(linha[1])
                nome = limpar_texto(linha[2])
                latitude = converter_numero_decimal(linha[3])
                longitude = converter_numero_decimal(linha[4])
                pais = detectar_pais_por_icao(icao)
                grupo_diaria = detectar_grupo_diaria_por_icao(icao)
                adicionar_aeroporto(aeroportos, iata, icao, nome, latitude, longitude, pais, grupo_diaria)

    print(f"Aeroportos carregados: {len(aeroportos)} códigos")
    return aeroportos


def limpar_texto(valor):
    if valor is None:
        return ""
    return str(valor).strip().upper()


def converter_numero_decimal(valor):
    if valor is None:
        return 0.0
    texto = str(valor).strip().replace(",", ".")
    if not texto:
        return 0.0
    try:
        return float(texto)
    except ValueError:
        return 0.0


def adicionar_aeroporto(aeroportos, iata, icao, nome, latitude, longitude, pais, grupo_diaria):
    if not iata and not icao:
        return
    if latitude == 0 or longitude == 0:
        return
    aeroporto = {"iata": iata, "icao": icao, "nome": nome, "lat": latitude, "lon": longitude, "pais": pais, "grupo_diaria": grupo_diaria}
    if iata:
        aeroportos[iata] = aeroporto
    if icao:
        aeroportos[icao] = aeroporto


def detectar_pais_por_icao(icao):
    icao = limpar_texto(icao)
    if icao.startswith("SB"):
        return "Brasil"
    if icao.startswith("SA"):
        return "Argentina"
    if icao.startswith("SC"):
        return "Chile"
    return "América do Sul"


def detectar_grupo_diaria_por_icao(icao):
    icao = limpar_texto(icao)
    if icao.startswith("SB"):
        return "NACIONAL"
    if icao.startswith("SA"):
        return "ARGENTINA"
    if icao.startswith("SC"):
        return "CHILE"
    return "AMERICA_DO_SUL"

# ============================================================
# LEITURA BRUTA / PERÍODO
# ============================================================

def extrair_linhas_brutas(sheet):
    rows = []
    for row in sheet.iter_rows(max_row=min(sheet.max_row, 120), values_only=True):
        converted = []
        has_content = False
        for cell in row:
            value = converter_valor_para_texto(cell)
            converted.append(value)
            if value.strip():
                has_content = True
        if has_content:
            rows.append(converted)
    return rows


def detectar_periodo_vigencia(sheet):
    limite_linhas = min(sheet.max_row, 15)
    limite_colunas = min(sheet.max_column, 25)
    for row in sheet.iter_rows(min_row=1, max_row=limite_linhas, min_col=1, max_col=limite_colunas, values_only=True):
        for cell in row:
            texto = converter_valor_para_texto(cell)
            periodo = tentar_extrair_periodo_do_texto(texto)
            if periodo:
                return periodo
    # fallback: usar menor/maior data encontrada nas primeiras colunas
    datas = []
    for row in sheet.iter_rows(values_only=True):
        for cell in list(row)[:6]:
            d = converter_para_data(cell)
            if d:
                datas.append(d)
    if datas:
        inicio = min(datas)
        fim = max(datas)
        return {"inicio": inicio, "fim": fim}
    return None


def tentar_extrair_periodo_do_texto(texto):
    if not texto:
        return None
    texto = str(texto)
    match = re.search(
        r"(\d{1,2})[-/ ]([A-Za-zÀ-ÿ]{3,})[-/ ](\d{2,4})\s*(?:to|a|até|ate|-)\s*(\d{1,2})[-/ ]([A-Za-zÀ-ÿ]{3,})[-/ ](\d{2,4})",
        texto,
        re.IGNORECASE,
    )
    if not match:
        return None
    dia_inicio = int(match.group(1))
    mes_inicio = converter_mes_texto_para_numero(match.group(2))
    ano_inicio = int(match.group(3))
    dia_fim = int(match.group(4))
    mes_fim = converter_mes_texto_para_numero(match.group(5))
    ano_fim = int(match.group(6))
    if ano_inicio < 100:
        ano_inicio += 2000
    if ano_fim < 100:
        ano_fim += 2000
    if not mes_inicio or not mes_fim:
        return None
    return {"inicio": date(ano_inicio, mes_inicio, dia_inicio), "fim": date(ano_fim, mes_fim, dia_fim)}


def data_dentro_periodo(data_evento, periodo):
    if periodo is None:
        return True
    return periodo["inicio"] <= data_evento <= periodo["fim"]

# ============================================================
# EXTRAÇÃO DA ESCALA
# ============================================================

def extrair_eventos_escala(sheet, periodo, aeroportos):
    eventos = []
    data_atual = None
    pairing_atual = ""
    jornada_atual = nova_jornada(None, "", "", "")

    for row in sheet.iter_rows(values_only=True):
        cells = list(row) + [""] * 40
        texto_linha = " ".join(converter_valor_para_texto(x).strip().upper() for x in cells if converter_valor_para_texto(x).strip())

        if not texto_linha:
            continue
        if eh_linha_cabecalho(texto_linha):
            continue

        # Colunas do formato que já estava funcionando.
        valor_data = cells[1]
        valor_pairing = cells[2]
        valor_duty_report = cells[4]
        valor_item = cells[5]
        valor_dep = cells[11]
        valor_arr = cells[12]
        valor_duty_debrief = cells[14]

        data_lida = converter_para_data(valor_data) or procurar_data_na_linha(cells)
        if data_lida:
            data_atual = data_lida

        if not data_atual:
            continue
        if not data_dentro_periodo(data_atual, periodo):
            continue

        pairing_texto = converter_valor_para_texto(valor_pairing).strip().upper()
        duty_report_texto = extrair_hora_texto(valor_duty_report) or extrair_hora_por_contexto(cells, ["DUTY REPORT", "REPORT", "APRESENTACAO", "APRESENTAÇÃO"])
        duty_debrief_texto = extrair_hora_texto(valor_duty_debrief) or extrair_hora_por_contexto(cells, ["DUTY DEBRIEF", "DEBRIEF"])

        if pairing_texto or duty_report_texto:
            if pairing_texto:
                pairing_atual = pairing_texto
            jornada_atual = nova_jornada(data_atual, pairing_atual, duty_report_texto, duty_debrief_texto)

        if duty_debrief_texto:
            jornada_atual["duty_debrief"] = duty_debrief_texto
            jornada_atual["duty_debrief_usado"] = False

        item_texto = converter_valor_para_texto(valor_item).strip().upper()
        if not item_texto:
            item_texto = detectar_item_na_linha(texto_linha)

        dep_texto = converter_valor_para_texto(valor_dep).strip().upper()
        arr_texto = converter_valor_para_texto(valor_arr).strip().upper()
        dep = extrair_aeroporto_hora(dep_texto)
        arr = extrair_aeroporto_hora(arr_texto)

        # Fallback para escalas com colunas deslocadas: pega os dois primeiros blocos AAA HH:MM da linha.
        if not dep or not arr:
            pares = extrair_todos_aeroportos_horas(texto_linha)
            if len(pares) >= 1 and not dep:
                dep = pares[0]
            if len(pares) >= 2 and not arr:
                arr = pares[-1]

        voo = extrair_codigo_voo(item_texto) or extrair_codigo_voo(texto_linha)
        servico = extrair_codigo_servico(item_texto, pairing_texto or pairing_atual or texto_linha)
        folga_codigo = detectar_folga(item_texto, pairing_texto, texto_linha)

        # DO/OFF/FOLGA deve aparecer na escala, mas não entra em salário, KM, diária ou horas.
        if folga_codigo:
            eventos.append(criar_evento_folga(data_atual, folga_codigo, pairing_atual, dep, arr))
            continue

        if voo and dep and arr:
            duty_report_evento, duty_debrief_evento = consumir_duty(jornada_atual, duty_debrief_texto)
            distancia_km = calcular_distancia_rota(dep["aeroporto"], arr["aeroporto"], aeroportos)
            km_periodos = dividir_km_por_periodo(data_atual, dep["hora"], arr["hora"], distancia_km)
            eventos.append({
                "data": data_atual.strftime("%d/%m/%Y"),
                "data_iso": data_atual.strftime("%Y-%m-%d"),
                "tipo": "VOO",
                "identificacao": voo,
                "pairing": jornada_atual.get("pairing", pairing_atual),
                "origem": dep["aeroporto"],
                "saida": dep["hora"],
                "destino": arr["aeroporto"],
                "chegada": arr["hora"],
                "duty_report": duty_report_evento,
                "duty_debrief": duty_debrief_evento,
                "duracao_horas": calcular_duracao_horas(data_atual, dep["hora"], arr["hora"]),
                "distancia_km": arredondar_2(distancia_km),
                "km_diurno": arredondar_2(km_periodos["km_diurno"]),
                "km_noturno": arredondar_2(km_periodos["km_noturno"]),
                "km_fim_semana": arredondar_2(km_periodos["km_fim_semana"]),
                "km_fim_semana_noturno": arredondar_2(km_periodos["km_fim_semana_noturno"]),
                "status": "OK" if distancia_km > 0 else "SEM DISTÂNCIA",
                "cafe": "", "almoco": "", "jantar": "", "ceia": "", "grupo_diaria": "", "moeda_diaria": "",
            })
            continue

        if servico:
            # Reserva/Sobreaviso podem vir com origem/horário; se não vier, ainda enviamos para aparecer na escala.
            duty_report_evento, duty_debrief_evento = consumir_duty(jornada_atual, duty_debrief_texto)
            origem = dep["aeroporto"] if dep else ""
            destino = arr["aeroporto"] if arr else origem
            saida = dep["hora"] if dep else duty_report_evento
            chegada = arr["hora"] if arr else duty_debrief_evento
            duracao = calcular_duracao_horas(data_atual, saida, chegada) if saida and chegada else 0
            eventos.append({
                "data": data_atual.strftime("%d/%m/%Y"),
                "data_iso": data_atual.strftime("%Y-%m-%d"),
                "tipo": servico["tipo"],
                "identificacao": servico["codigo"],
                "pairing": jornada_atual.get("pairing", pairing_atual),
                "origem": origem,
                "saida": saida,
                "destino": destino,
                "chegada": chegada,
                "duty_report": duty_report_evento,
                "duty_debrief": duty_debrief_evento,
                "duracao_horas": arredondar_2(duracao),
                "distancia_km": 0,
                "km_diurno": 0,
                "km_noturno": 0,
                "km_fim_semana": 0,
                "km_fim_semana_noturno": 0,
                "status": "OK",
                "cafe": "", "almoco": "", "jantar": "", "ceia": "", "grupo_diaria": "", "moeda_diaria": "",
            })
            continue

    eventos = ordenar_eventos(eventos)
    eventos = remover_folgas_em_dias_com_atividade(eventos)
    eventos = deduplicar_eventos_escala(eventos)
    eventos = preencher_folgas_dias_sem_evento(eventos, periodo)
    eventos = remover_folgas_em_dias_com_atividade(eventos)
    eventos = deduplicar_eventos_escala(eventos)
    return ordenar_eventos(eventos)


def nova_jornada(data_atual, pairing, duty_report, duty_debrief):
    return {
        "data": data_atual,
        "pairing": pairing or "",
        "duty_report": duty_report or "",
        "duty_debrief": duty_debrief or "",
        "duty_report_usado": False,
        "duty_debrief_usado": False,
    }


def eh_linha_cabecalho(texto_linha):
    cabecalhos = ["DUTY REPORT", "DEP STN", "ARR STN", "DEPARTURE", "ARRIVAL", "ITEM", "PAIRING"]
    # Não pule uma linha que contenha um evento real junto, como ASB/HSB/LA/DO.
    tem_evento = bool(extrair_codigo_voo(texto_linha) or re.search(r"\b(ASB\d*|HSB|HSBI|HSBE|DO|OFF)\b", texto_linha))
    return any(c in texto_linha for c in cabecalhos) and not tem_evento


def procurar_data_na_linha(cells):
    for cell in list(cells)[:8]:
        d = converter_para_data(cell)
        if d:
            return d
    return None


def detectar_item_na_linha(texto):
    voo = extrair_codigo_voo(texto)
    if voo:
        return voo
    for pattern in [r"\bHSBE\b", r"\bHSBI\b", r"\bHSB\b", r"\bASB\s*\d*\b", r"\bDO\b", r"\bOFF\b", r"\bFOLGA\b", r"\bDAY OFF\b", r"\bD/O\b"]:
        m = re.search(pattern, texto.upper())
        if m:
            return m.group(0).replace(" ", "")
    return ""


def detectar_folga(item_texto, pairing_texto, texto_linha):
    texto = f" {item_texto or ''} {pairing_texto or ''} {texto_linha or ''} ".upper()
    if re.search(r"\b(DO|D/O|OFF|FOLGA|DAY OFF|DAYOFF)\b", texto):
        # Evita confundir palavras que contenham DO dentro de outro texto.
        if re.search(r"\bDO\b", texto):
            return "DO"
        if re.search(r"\bD/O\b", texto):
            return "DO"
        if re.search(r"\bOFF\b", texto):
            return "OFF"
        if re.search(r"\bFOLGA\b", texto):
            return "FOLGA"
        if re.search(r"\bDAY OFF\b|\bDAYOFF\b", texto):
            return "OFF"
    return ""


def criar_evento_folga(data_atual, codigo, pairing_atual, dep, arr):
    origem = dep["aeroporto"] if dep else ""
    destino = arr["aeroporto"] if arr else origem
    saida = dep["hora"] if dep else "00:00"
    chegada = arr["hora"] if arr else "23:59"
    return {
        "data": data_atual.strftime("%d/%m/%Y"),
        "data_iso": data_atual.strftime("%Y-%m-%d"),
        "tipo": "FOLGA",
        "identificacao": codigo or "DO",
        "pairing": pairing_atual or codigo or "DO",
        "origem": origem,
        "saida": saida,
        "destino": destino,
        "chegada": chegada,
        "duty_report": "",
        "duty_debrief": "",
        "duracao_horas": 0,
        "distancia_km": 0,
        "km_diurno": 0,
        "km_noturno": 0,
        "km_fim_semana": 0,
        "km_fim_semana_noturno": 0,
        "status": "OK",
        "cafe": "", "almoco": "", "jantar": "", "ceia": "", "grupo_diaria": "", "moeda_diaria": "",
    }


def consumir_duty(jornada_atual, duty_debrief_texto):
    duty_report_evento = ""
    duty_debrief_evento = ""
    if not jornada_atual.get("duty_report_usado"):
        duty_report_evento = jornada_atual.get("duty_report", "")
        jornada_atual["duty_report_usado"] = True
    if duty_debrief_texto and not jornada_atual.get("duty_debrief_usado"):
        duty_debrief_evento = duty_debrief_texto
        jornada_atual["duty_debrief_usado"] = True
    return duty_report_evento, duty_debrief_evento


def ordenar_eventos(eventos):
    def chave(event):
        d = converter_para_data(event.get("data_iso")) or date(1900, 1, 1)
        h = event.get("saida") or event.get("duty_report") or "00:00"
        try:
            dt = criar_datetime(d, h)
        except Exception:
            dt = datetime.combine(d, time(0, 0))
        # Folga fica depois de eventos reais do dia se houver.
        ordem_tipo = 9 if event.get("tipo") == "FOLGA" else 1
        return (dt, ordem_tipo)
    return sorted(eventos, key=chave)


def preencher_folgas_dias_sem_evento(eventos, periodo):
    if periodo is None:
        return eventos

    # Só cria folga automática quando o dia não possui nenhum evento real.
    # Isso evita o xabu de aparecer DO/OFF no mesmo dia de HSB/ASB/voo.
    datas_com_evento = {e.get("data_iso") for e in eventos if e.get("data_iso")}
    novos = list(eventos)
    cursor = periodo["inicio"]

    while cursor <= periodo["fim"]:
        iso = cursor.strftime("%Y-%m-%d")
        if iso not in datas_com_evento:
            novos.append(criar_evento_folga(cursor, "DO", "DO", None, None))
        cursor += timedelta(days=1)

    return ordenar_eventos(novos)


def remover_folgas_em_dias_com_atividade(eventos):
    """Remove DO/OFF de dias que já possuem atividade real.

    A escala LATAM às vezes traz linhas auxiliares DO/OFF no mesmo dia de HSB ou
    outras atividades. Para a visualização principal, folga só deve aparecer em
    dia realmente livre.
    """
    datas_com_atividade = {
        e.get("data_iso")
        for e in eventos
        if e.get("data_iso") and e.get("tipo") != "FOLGA"
    }

    return [
        e for e in eventos
        if not (e.get("tipo") == "FOLGA" and e.get("data_iso") in datas_com_atividade)
    ]


def deduplicar_eventos_escala(eventos):
    vistos = set()
    resultado = []

    for e in eventos:
        chave = (
            e.get("data_iso", ""),
            e.get("tipo", ""),
            e.get("identificacao", ""),
            e.get("origem", ""),
            e.get("saida", ""),
            e.get("destino", ""),
            e.get("chegada", ""),
        )

        if chave in vistos:
            continue

        vistos.add(chave)
        resultado.append(e)

    return ordenar_eventos(resultado)

# ============================================================
# RESUMO / DIÁRIAS / HOLERITE
# ============================================================

def calcular_resumo_inicial(events, aeroportos, cargo):
    total_voos = total_reservas = total_sobreavisos = 0
    horas_reserva = horas_sobreaviso = 0.0
    km_total = km_diurno = km_noturno = km_fim_semana = km_fim_semana_noturno = 0.0
    voos_sem_distancia = []

    for event in events:
        tipo = event.get("tipo", "")
        duracao = float(event.get("duracao_horas") or 0)
        if tipo == "VOO":
            total_voos += 1
            distancia = float(event.get("distancia_km") or 0)
            km_total += distancia
            km_diurno += float(event.get("km_diurno") or 0)
            km_noturno += float(event.get("km_noturno") or 0)
            km_fim_semana += float(event.get("km_fim_semana") or 0)
            km_fim_semana_noturno += float(event.get("km_fim_semana_noturno") or 0)
            if distancia <= 0:
                voos_sem_distancia.append(f'{event.get("origem", "")}-{event.get("destino", "")}')
        elif tipo == "RESERVA":
            total_reservas += 1
            horas_reserva += duracao
        elif tipo == "SOBREAVISO":
            total_sobreavisos += 1
            horas_sobreaviso += duracao

    diarias = calcular_diarias(events, aeroportos)
    resumo_base = {
        "total_eventos": len(events),
        "total_voos": total_voos,
        "total_reservas": total_reservas,
        "total_sobreavisos": total_sobreavisos,
        "horas_reserva": arredondar_2(horas_reserva),
        "horas_sobreaviso": arredondar_2(horas_sobreaviso),
        "km_total": round(km_total),
        "km_diurno": round(km_diurno),
        "km_noturno": round(km_noturno),
        "km_fim_semana": round(km_fim_semana),
        "km_fim_semana_noturno": round(km_fim_semana_noturno),
        "voos_sem_distancia": sorted(list(set(voos_sem_distancia))),
        "diarias": diarias["resumo"],
        "total_diarias_brl": diarias["total_brl"],
        "total_diarias_usd": diarias["total_usd"],
    }
    resumo_base["holerite"] = calcular_holerite(resumo_base, cargo)
    return resumo_base


def calcular_diarias(events, aeroportos):
    limpar_marcacoes_diarias(events)
    jornadas = marcar_diarias_nas_linhas_da_escala(events, aeroportos)
    resumo = criar_estrutura_diarias()
    total_brl = total_usd = 0.0

    for jornada in jornadas:
        grupo = jornada["grupo_diaria"]
        moeda = obter_moeda_grupo(grupo)
        valor_total_jornada = 0.0
        for refeicao, marcado in jornada["refeicoes"].items():
            if not marcado:
                continue
            valor_refeicao = calcular_valor_refeicao(grupo, refeicao, 1)
            resumo[grupo][refeicao]["quantidade"] += 1
            resumo[grupo][refeicao]["valor_total"] += valor_refeicao
            resumo[grupo]["total"] += valor_refeicao
            valor_total_jornada += valor_refeicao
        if moeda == "BRL":
            total_brl += valor_total_jornada
        else:
            total_usd += valor_total_jornada

    calcular_valores_unitarios_diarias(resumo)
    resumo["total_brl"] = arredondar_2(total_brl)
    resumo["total_usd"] = arredondar_2(total_usd)
    resumo["total_geral"] = arredondar_2(total_brl)
    return {"resumo": resumo, "total_brl": arredondar_2(total_brl), "total_usd": arredondar_2(total_usd)}


def limpar_marcacoes_diarias(events):
    for event in events:
        event["cafe"] = ""
        event["almoco"] = ""
        event["jantar"] = ""
        event["ceia"] = ""
        event["grupo_diaria"] = ""
        event["moeda_diaria"] = ""


def marcar_diarias_nas_linhas_da_escala(events, aeroportos):
    jornadas = []
    jornada_atual = None
    numero_jornada = 1
    for index, event in enumerate(events):
        tipo = event.get("tipo", "")
        if tipo in ["SOBREAVISO", "FOLGA"]:
            continue
        data_base = converter_para_data(event.get("data_iso"))
        if data_base is None:
            continue
        duty_report = event.get("duty_report", "")
        duty_debrief = event.get("duty_debrief", "")
        origem = event.get("origem", "")
        destino = event.get("destino", "")
        if duty_report:
            if jornada_atual is not None:
                fechar_jornada_sem_debrief(jornada_atual, aeroportos)
                finalizar_jornada_diaria(jornada_atual, events, aeroportos)
                jornadas.append(jornada_atual)
                numero_jornada += 1
            inicio = criar_datetime(data_base, duty_report)
            jornada_atual = {
                "numero": numero_jornada,
                "inicio": inicio,
                "fim": None,
                "linha_marcacao": index,
                "aeroportos": set(),
                "ultimo_horario": inicio,
                "grupo_diaria": "NACIONAL",
                "moeda": "BRL",
                "refeicoes": {"cafe": False, "almoco": False, "jantar": False, "ceia": False},
            }
        if jornada_atual is None:
            continue
        if origem:
            jornada_atual["aeroportos"].add(origem)
        if destino:
            jornada_atual["aeroportos"].add(destino)
        chegada = event.get("chegada", "")
        if chegada:
            chegada_dt = criar_datetime(data_base, chegada)
            if chegada_dt < jornada_atual["inicio"]:
                chegada_dt += timedelta(days=1)
            jornada_atual["ultimo_horario"] = chegada_dt
        if duty_debrief:
            fim = criar_datetime(data_base, duty_debrief)
            if fim < jornada_atual["inicio"]:
                fim += timedelta(days=1)
            jornada_atual["fim"] = fim
            finalizar_jornada_diaria(jornada_atual, events, aeroportos)
            jornadas.append(jornada_atual)
            jornada_atual = None
            numero_jornada += 1
    if jornada_atual is not None:
        fechar_jornada_sem_debrief(jornada_atual, aeroportos)
        finalizar_jornada_diaria(jornada_atual, events, aeroportos)
        jornadas.append(jornada_atual)
    return jornadas


def fechar_jornada_sem_debrief(jornada, aeroportos):
    if jornada.get("fim") is None:
        jornada["fim"] = jornada.get("ultimo_horario") or jornada.get("inicio")
    jornada["grupo_diaria"] = detectar_grupo_diaria_da_jornada(jornada.get("aeroportos", set()), aeroportos)
    jornada["moeda"] = obter_moeda_grupo(jornada["grupo_diaria"])
    jornada["refeicoes"] = calcular_refeicoes_da_jornada(jornada["inicio"], jornada["fim"])


def finalizar_jornada_diaria(jornada, events, aeroportos):
    jornada["grupo_diaria"] = detectar_grupo_diaria_da_jornada(jornada.get("aeroportos", set()), aeroportos)
    jornada["moeda"] = obter_moeda_grupo(jornada["grupo_diaria"])
    jornada["refeicoes"] = calcular_refeicoes_da_jornada(jornada["inicio"], jornada["fim"])
    linha = jornada.get("linha_marcacao")
    if linha is None or linha < 0 or linha >= len(events):
        return
    events[linha]["cafe"] = "SIM" if jornada["refeicoes"]["cafe"] else ""
    events[linha]["almoco"] = "SIM" if jornada["refeicoes"]["almoco"] else ""
    events[linha]["jantar"] = "SIM" if jornada["refeicoes"]["jantar"] else ""
    events[linha]["ceia"] = "SIM" if jornada["refeicoes"]["ceia"] else ""
    events[linha]["grupo_diaria"] = jornada["grupo_diaria"]
    events[linha]["moeda_diaria"] = jornada["moeda"]


def detectar_grupo_diaria_da_jornada(codigos_aeroportos, aeroportos):
    grupos_encontrados = set()
    for codigo in codigos_aeroportos:
        aeroporto = aeroportos.get(str(codigo).strip().upper())
        if not aeroporto:
            continue
        grupo = aeroporto.get("grupo_diaria", "")
        if grupo:
            grupos_encontrados.add(grupo)
    if "CHILE" in grupos_encontrados:
        return "CHILE"
    if "ARGENTINA" in grupos_encontrados:
        return "ARGENTINA"
    if "AMERICA_DO_SUL" in grupos_encontrados:
        return "AMERICA_DO_SUL"
    return "NACIONAL"


def calcular_refeicoes_da_jornada(inicio, fim):
    refeicoes = {"cafe": False, "almoco": False, "jantar": False, "ceia": False}
    if inicio is None or fim is None:
        return refeicoes
    if fim < inicio:
        fim += timedelta(days=1)
    data_cursor = inicio.date()
    while data_cursor <= fim.date():
        janelas = {
            "cafe": (datetime.combine(data_cursor, time(5, 0)), datetime.combine(data_cursor, time(8, 0))),
            "almoco": (datetime.combine(data_cursor, time(11, 0)), datetime.combine(data_cursor, time(13, 0))),
            "jantar": (datetime.combine(data_cursor, time(19, 0)), datetime.combine(data_cursor, time(20, 0))),
            "ceia": (datetime.combine(data_cursor, time(0, 0)), datetime.combine(data_cursor, time(1, 0))),
        }
        for refeicao, janela in janelas.items():
            if intervalo_intersecta_inclusivo(inicio, fim, janela[0], janela[1]):
                refeicoes[refeicao] = True
        data_cursor += timedelta(days=1)
    return refeicoes


def intervalo_intersecta_inclusivo(inicio_a, fim_a, inicio_b, fim_b):
    return inicio_a <= fim_b and fim_a >= inicio_b


def criar_estrutura_diarias():
    grupos = ["NACIONAL", "ARGENTINA", "CHILE", "AMERICA_DO_SUL"]
    refeicoes = ["cafe", "almoco", "jantar", "ceia"]
    resultado = {}
    for grupo in grupos:
        resultado[grupo] = {"moeda": obter_moeda_grupo(grupo)}
        for refeicao in refeicoes:
            resultado[grupo][refeicao] = {"quantidade": 0, "valor_unitario": 0, "valor_total": 0}
        resultado[grupo]["total"] = 0
    resultado["total_brl"] = 0
    resultado["total_usd"] = 0
    resultado["total_geral"] = 0
    return resultado


def obter_moeda_grupo(grupo):
    return "BRL" if grupo == "NACIONAL" else "USD"


def calcular_valor_refeicao(grupo, refeicao, quantidade):
    valor_principal = obter_valor_refeicao_principal(grupo)
    valor_unitario = valor_principal * 0.25 if refeicao == "cafe" else valor_principal
    return quantidade * valor_unitario


def obter_valor_refeicao_principal(grupo):
    valores = {"NACIONAL": 105.04, "ARGENTINA": 22.05, "CHILE": 25.15, "AMERICA_DO_SUL": 21.00}
    return valores.get(grupo, 0)


def calcular_valores_unitarios_diarias(resultado):
    for grupo, dados_grupo in resultado.items():
        if grupo in ["total_brl", "total_usd", "total_geral"]:
            continue
        for refeicao in ["cafe", "almoco", "jantar", "ceia"]:
            valor_unitario = obter_valor_refeicao_principal(grupo)
            if refeicao == "cafe":
                valor_unitario *= 0.25
            dados_grupo[refeicao]["valor_unitario"] = arredondar_2(valor_unitario)
            dados_grupo[refeicao]["valor_total"] = arredondar_2(dados_grupo[refeicao]["valor_total"])
        dados_grupo["total"] = arredondar_2(dados_grupo["total"])

# ============================================================
# HOLERITE
# ============================================================

def obter_config_cargo(cargo):
    cargo = limpar_texto(cargo)
    configs = {
        "COPILOTO": {"cargo": "COPILOTO", "salario_base": 9732.85, "km_diurno": 0.143193, "km_noturno": 0.286386, "km_fim_semana": 0.286386, "km_fim_semana_noturno": 0.286386, "hora_reserva": 121.71, "hora_sobreaviso": 40.57, "hora_simulador": 508.45, "gratificacao": 0.0, "gratificacao_ativa": True, "assistencia_medica_amil": 443.05, "repouso_percentual_sobre_variaveis": 0.363636},
        "COMANDANTE": {"cargo": "COMANDANTE", "salario_base": 16727.06, "km_diurno": 0.216027, "km_noturno": 0.432054, "km_fim_semana": 0.432054, "km_fim_semana_noturno": 0.432054, "hora_reserva": 183.61, "hora_sobreaviso": 61.20, "hora_simulador": 753.61, "gratificacao": 5018.12, "gratificacao_ativa": True, "assistencia_medica_amil": 443.05, "repouso_percentual_sobre_variaveis": 0.363636},
        "COMISSARIO": {"cargo": "COMISSARIO", "salario_base": 3013.65, "km_diurno": 0.057349, "km_noturno": 0.114698, "km_fim_semana": 0.114698, "km_fim_semana_noturno": 0.114698, "hora_reserva": 48.75, "hora_sobreaviso": 16.25, "hora_simulador": 0.0, "gratificacao": 0.0, "gratificacao_ativa": True, "assistencia_medica_amil": 443.05, "repouso_percentual_sobre_variaveis": 0.363636},
    }
    base = configs.get(cargo, configs["COPILOTO"])
    base.setdefault("previdencia_privada_percentual", 0.0)
    base.setdefault("servico_saude_dasa", 0.0)
    base.setdefault("seguro_vida_bradesco_funeral", 0.0)
    base.setdefault("seguro_vida_complementar", 0.0)
    base.setdefault("assistencia_odonto_familia", 0.0)
    base.setdefault("gympass", 0.0)
    return base


def calcular_holerite(summary, cargo):
    valores = obter_config_cargo(cargo)
    quantidade_km_diurno = float(summary.get("km_diurno") or 0)
    quantidade_km_noturno = float(summary.get("km_noturno") or 0)
    quantidade_km_fim_semana = float(summary.get("km_fim_semana") or 0)
    quantidade_km_fim_semana_noturno = float(summary.get("km_fim_semana_noturno") or 0)
    horas_reserva = float(summary.get("horas_reserva") or 0)
    horas_sobreaviso = float(summary.get("horas_sobreaviso") or 0)

    salario_base = valores["salario_base"]
    km_diurno_valor = quantidade_km_diurno * valores["km_diurno"]
    km_noturno_valor = quantidade_km_noturno * valores["km_noturno"]
    km_fim_semana_valor = quantidade_km_fim_semana * valores["km_fim_semana"]
    km_fim_semana_noturno_valor = quantidade_km_fim_semana_noturno * valores["km_fim_semana_noturno"]
    reserva_valor = horas_reserva * valores["hora_reserva"]
    sobreaviso_valor = horas_sobreaviso * valores["hora_sobreaviso"]
    simulador_valor = 0.0
    variaveis_para_repouso = km_diurno_valor + km_noturno_valor + km_fim_semana_valor + km_fim_semana_noturno_valor + reserva_valor + sobreaviso_valor + simulador_valor
    repouso_remunerado = variaveis_para_repouso * valores["repouso_percentual_sobre_variaveis"]
    gratificacao_valor = valores["gratificacao"] if valores["gratificacao_ativa"] else 0.0

    proventos = [
        criar_linha_provento("Salario Base", 1, salario_base, salario_base),
        criar_linha_provento("KM Diurno", quantidade_km_diurno, valores["km_diurno"], km_diurno_valor),
        criar_linha_provento("KM Noturno", quantidade_km_noturno, valores["km_noturno"], km_noturno_valor),
        criar_linha_provento("KM Fim de Semana", quantidade_km_fim_semana, valores["km_fim_semana"], km_fim_semana_valor),
        criar_linha_provento("KM Fim de Semana NOT", quantidade_km_fim_semana_noturno, valores["km_fim_semana_noturno"], km_fim_semana_noturno_valor),
        criar_linha_provento("Horas Reserva", horas_reserva, valores["hora_reserva"], reserva_valor),
        criar_linha_provento("Sobreaviso", horas_sobreaviso, valores["hora_sobreaviso"], sobreaviso_valor),
        criar_linha_provento("Simulador", 0, valores["hora_simulador"], simulador_valor),
        criar_linha_provento("Repouso Remunerado", "", "", repouso_remunerado),
        criar_linha_provento("Gratificação", "", "", gratificacao_valor),
    ]
    total_proventos = sum(float(linha["final"] or 0) for linha in proventos)
    inss_remuneracao = calcular_inss_remuneracao(total_proventos)
    base_ir = total_proventos - inss_remuneracao
    irrf_salario = calcular_irrf_salario(base_ir)
    descontos = [
        criar_linha_desconto("Previdencia Privada", "0%", 0.0),
        criar_linha_desconto("Assistencia Medica AMIL", "1", valores["assistencia_medica_amil"]),
        criar_linha_desconto("Servico de Saude DASA", False, valores.get("servico_saude_dasa", 0.0)),
        criar_linha_desconto("Seguro de Vida Bradesco Funeral", False, valores.get("seguro_vida_bradesco_funeral", 0.0)),
        criar_linha_desconto("Seguro de Vida Complementar", "Nao Utilizo", valores.get("seguro_vida_complementar", 0.0)),
        criar_linha_desconto("Assistencia Odonto Familia", "Nao Utilizo", valores.get("assistencia_odonto_familia", 0.0)),
        criar_linha_desconto("Gympass", "Nao Utilizo", valores.get("gympass", 0.0)),
        criar_linha_desconto("IRRF salario", "", irrf_salario),
    ]
    desconto_total = inss_remuneracao + sum(float(linha["valor"] or 0) for linha in descontos)
    salario_liquido = total_proventos - desconto_total
    return {
        "cargo": valores["cargo"],
        "proventos": proventos,
        "base_ir": {"total_proventos": arredondar_2(total_proventos), "inss_remuneracao": arredondar_2(inss_remuneracao), "base_ir": arredondar_2(base_ir)},
        "descontos": descontos,
        "salario": {"proventos": arredondar_2(total_proventos), "descontos": arredondar_2(desconto_total), "salario_liquido": arredondar_2(salario_liquido)},
    }


def criar_linha_provento(descricao, quantidade, razao, final):
    return {"descricao": descricao, "quantidade": quantidade, "razao": razao, "final": arredondar_2(final)}


def criar_linha_desconto(descricao, opcao, valor):
    return {"descricao": descricao, "opcao": opcao, "valor": arredondar_2(valor)}


def calcular_inss_remuneracao(total_proventos):
    if total_proventos <= 0:
        return 0.0
    return min(total_proventos * 0.14, 988.07)


def calcular_irrf_salario(base_ir):
    if base_ir <= 2428.80:
        return 0.0
    if base_ir <= 2826.65:
        return (base_ir * 0.075) - 182.16
    if base_ir <= 3751.05:
        return (base_ir * 0.15) - 394.16
    if base_ir <= 4664.68:
        return (base_ir * 0.225) - 675.49
    return (base_ir * 0.275) - 908.73

# ============================================================
# DISTÂNCIA / KM
# ============================================================

def calcular_distancia_rota(origem, destino, aeroportos):
    origem = str(origem).strip().upper()
    destino = str(destino).strip().upper()
    if origem not in aeroportos or destino not in aeroportos:
        return 0.0
    a = aeroportos[origem]
    b = aeroportos[destino]
    return haversine_km(a["lat"], a["lon"], b["lat"], b["lon"])


def haversine_km(lat1, lon1, lat2, lon2):
    raio_terra_km = 6371.0
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    h = math.sin(delta_phi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2) ** 2
    c = 2 * math.atan2(math.sqrt(h), math.sqrt(1 - h))
    return raio_terra_km * c


def dividir_km_por_periodo(data_base, hora_saida, hora_chegada, distancia_km):
    if distancia_km <= 0:
        return {"km_diurno": 0, "km_noturno": 0, "km_fim_semana": 0, "km_fim_semana_noturno": 0}
    inicio = criar_datetime(data_base, hora_saida)
    fim = criar_datetime(data_base, hora_chegada)
    if fim <= inicio:
        fim += timedelta(days=1)
    duracao_minutos = int((fim - inicio).total_seconds() / 60)
    if duracao_minutos <= 0:
        return {"km_diurno": 0, "km_noturno": 0, "km_fim_semana": 0, "km_fim_semana_noturno": 0}
    km_por_minuto = distancia_km / duracao_minutos
    km_diurno = km_noturno = km_fim_semana = km_fim_semana_noturno = 0.0
    momento = inicio
    for _ in range(duracao_minutos):
        hora_decimal = momento.hour + momento.minute / 60
        noturno = hora_decimal >= 18 or hora_decimal < 6
        fim_semana = momento.weekday() in [5, 6]
        if fim_semana and noturno:
            km_fim_semana_noturno += km_por_minuto
        elif fim_semana:
            km_fim_semana += km_por_minuto
        elif noturno:
            km_noturno += km_por_minuto
        else:
            km_diurno += km_por_minuto
        momento += timedelta(minutes=1)
    return {"km_diurno": km_diurno, "km_noturno": km_noturno, "km_fim_semana": km_fim_semana, "km_fim_semana_noturno": km_fim_semana_noturno}

# ============================================================
# CONVERSÕES / PARSERS
# ============================================================

def converter_valor_para_texto(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return str(value)


def converter_para_data(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    texto = str(value).strip()
    if not texto:
        return None
    formatos = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d-%b-%Y", "%d-%B-%Y", "%Y-%m-%d %H:%M:%S"]
    for formato in formatos:
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            pass
    match = re.search(r"(\d{1,2})[-/ ]([A-Za-zÀ-ÿ]{3,})[-/ ](\d{2,4})", texto)
    if match:
        dia = int(match.group(1))
        mes = converter_mes_texto_para_numero(match.group(2))
        ano = int(match.group(3))
        if ano < 100:
            ano += 2000
        if mes:
            return date(ano, mes, dia)
    return None


def converter_mes_texto_para_numero(mes_texto):
    meses = {
        "JAN": 1, "JANEIRO": 1, "JANUARY": 1,
        "FEV": 2, "FEB": 2, "FEVEREIRO": 2, "FEBRUARY": 2,
        "MAR": 3, "MARCO": 3, "MARÇO": 3, "MARCH": 3,
        "ABR": 4, "APR": 4, "ABRIL": 4, "APRIL": 4,
        "MAI": 5, "MAY": 5, "MAIO": 5,
        "JUN": 6, "JUNE": 6, "JUNHO": 6,
        "JUL": 7, "JULY": 7, "JULHO": 7,
        "AGO": 8, "AUG": 8, "AGOSTO": 8, "AUGUST": 8,
        "SET": 9, "SEP": 9, "SEPT": 9, "SETEMBRO": 9, "SEPTEMBER": 9,
        "OUT": 10, "OCT": 10, "OUTUBRO": 10, "OCTOBER": 10,
        "NOV": 11, "NOVEMBRO": 11, "NOVEMBER": 11,
        "DEZ": 12, "DEC": 12, "DEZEMBRO": 12, "DECEMBER": 12,
    }
    chave = str(mes_texto).strip().upper().replace(".", "")
    return meses.get(chave)


def extrair_hora_texto(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    texto = str(value).strip()
    match = re.search(r"\b(\d{1,2}:\d{2})(\(\+1\))?\b", texto)
    if match:
        return match.group(0)
    return ""


def extrair_hora_por_contexto(cells, chaves):
    textos = [converter_valor_para_texto(c).strip().upper() for c in cells]
    for idx, texto in enumerate(textos):
        if any(chave in texto for chave in chaves):
            for j in range(max(0, idx - 2), min(len(textos), idx + 4)):
                hora = extrair_hora_texto(textos[j])
                if hora:
                    return hora
    return ""


def extrair_codigo_voo(texto):
    if not texto:
        return None
    match = re.search(r"\b[A-Z]{2}\d{3,4}\b", texto.upper())
    return match.group(0) if match else None


def extrair_codigo_servico(item_texto, pairing_texto):
    texto = f" {item_texto or ''} {pairing_texto or ''} ".upper()
    match_sobreaviso = re.search(r"\b(HSBE|HSBI|HSB)\b", texto)
    if match_sobreaviso:
        return {"tipo": "SOBREAVISO", "codigo": match_sobreaviso.group(1)}
    match_reserva = re.search(r"\bASB\s*\d*\b", texto)
    if match_reserva:
        return {"tipo": "RESERVA", "codigo": match_reserva.group(0).replace(" ", "")}
    return None


def extrair_aeroporto_hora(texto):
    if not texto:
        return None
    match = re.search(r"\b([A-Z]{3})\s+(\d{1,2}:\d{2}(?:\(\+1\))?)\b", texto.upper())
    if not match:
        return None
    return {"aeroporto": match.group(1), "hora": match.group(2)}


def extrair_todos_aeroportos_horas(texto):
    pares = []
    for match in re.finditer(r"\b([A-Z]{3})\s+(\d{1,2}:\d{2}(?:\(\+1\))?)\b", texto.upper()):
        pares.append({"aeroporto": match.group(1), "hora": match.group(2)})
    return pares


def calcular_duracao_horas(data_base, hora_inicio, hora_fim):
    if not hora_inicio or not hora_fim:
        return 0.0
    inicio = criar_datetime(data_base, hora_inicio)
    fim = criar_datetime(data_base, hora_fim)
    if fim <= inicio:
        fim += timedelta(days=1)
    duracao = (fim - inicio).total_seconds() / 3600
    return arredondar_2(duracao)


def criar_datetime(data_base, hora_texto):
    hora_limpa = str(hora_texto).replace("(+1)", "").strip()
    partes = hora_limpa.split(":")
    dt = datetime(data_base.year, data_base.month, data_base.day, int(partes[0]), int(partes[1]), 0)
    if "(+1)" in str(hora_texto):
        dt += timedelta(days=1)
    return dt


def arredondar_2(valor):
    if valor == "":
        return ""
    try:
        return round(float(valor), 2)
    except (TypeError, ValueError):
        return 0.0
